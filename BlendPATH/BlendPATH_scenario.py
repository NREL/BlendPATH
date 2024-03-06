from dataclasses import dataclass,field
from ctypes import *
import numpy as np
import pandas as pd
import BlendPATH.Design_assessment_functions as daf
import BlendPATH.modification_functions as mod_fxn
import shutil
import openpyxl
import ProFAST
import os
import json
from pkg_resources import resource_filename

def run_BlendPATH_file(filename,batch_id=None):
    input_csv = read_input_file(filename)

    financial_cols = ['analysis_start_year','operating_life','installation_months','long_term_utilization','property_tax_insurance',
                            'admin_expense','total_income_tax_rate','capital_gains_tax_rate','general_inflation','discount_rate',
                            'debt_equity_ratio','debt_type','loan_period','interest_rate','cash_onhand']

    if not batch_id == None:
        input_csv = input_csv.loc[input_csv['index'].isin(batch_id)]
    for index,row in input_csv.iterrows():
        blendpath = BlendPATH_scenario(casestudy_name = row['casestudy_name'],
                                save_directory = row['save_directory'],
                                design_option = row['design_option'],
                                location_class  = row['location_class'],
                                joint_factor  = row['joint_factor'],
                                design_pressure_known = row['design_pressure_known'],
                                pressure_drop_factor  = row['pressure_drop_factor'],
                                compression_ratio  = row['compression_ratio'],
                                blending = row['blending'],
                                natural_gas_cost = row['natural_gas_cost'],
                                hydrogen_cost = row['hydrogen_cost'],
                                nodes  = row['nodes'],
                                final_outlet_pressure=row['final_outlet_pressure'],
                                region=row['region'],
                                batch_id = row['index'])
        
        # Check if material cost is overwritten
        if not pd.isna(row['OVERRIDE: material ($/kg)']):
            try:
                override_mat = json.loads(row['OVERRIDE: material ($/kg)'].replace("“","\"").replace("”","\"")) 
            except:
                raise Exception('OVERRIDE: material ($/kg) is formatted poorly. Please use the {"X70":1,"X65":0.5} format or leave empty')
            for grade,price in override_mat.items():
                blendpath.steel_costs_per_kg.loc[blendpath.steel_costs_per_kg['Grade']==grade,'Price [$/kg]'] = price

        
        blendpath.run_mod(mod_type=row['modification'],design_option=row['modification_design_option'])

        cost_overrides = row[['OVERRIDE: Original pipeline cost ($)','OVERRIDE: labor cost ($/in/mi)','OVERRIDE: row cost ($/in/mi)','OVERRIDE: misc cost ($/in/mi)']]
        LCOH_i = blendpath.run_financial(cost_overrides,row[financial_cols].to_dict())

def read_input_file(filename):
    # Check if file exists
    if not os.path.isfile(filename):
        raise Exception(f'Cannot find file with name {filename}')

    # Read csv into pandas df
    input_csv = pd.read_csv(filename)

    # Check entries for wrong modification method. Need to extend this to other columns
    check_mod = input_csv['modification'].isin(['parallel_loop','additional_compressors','direct_replacement'])
    if not check_mod.any():
        raise Exception(f'modification must be either parallel_loop, additional_compressors, or direct_replacement. Check entry {check_mod.loc[check_mod==False].index.tolist()}')

    return input_csv

def disclaimer_message():
    return '\n---------DISCLAIMER---------\nAS NOTED IN THE LICENSE, ALLIANCE FOR SUSTAINABLE ENERGY, LLC: (i) DISCLAIMS ANY WARRANTIES, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO ANY IMPLIED WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE, TITLE OR NON-INFRINGEMENT, AND (ii) DOES NOT ASSUME ANY LEGAL LIABILITY OR RESPONSIBILITY FOR THE ACCURACY, COMPLETENESS, OR USEFULNESS OF SOFTWARE OR ITS OUTPUTS. ANY RELIANCE BY THE USER ON THE SOFTWARE IS DONE AT THE USER’S OWN RISK.\n\nFurther, BlendPATH was developed to provide the user with economic analytical capabilities to evaluate potential natural gas transmission pipeline modifications for accommodating hydrogen blending and to assess the potential associated economic impacts. This tool is intended for use during the early stages of hydrogen blending concept evaluation, which is defined here as the initial project phase when the developer has already gathered the relevant pipeline material design and operation data but has not yet conducted the detailed pipeline material testing, as specified in ASME B31.12, or performed inline inspections that qualify a given pipeline section for hydrogen blending. Despite BlendPATH utilizing ASME B31.12 to assess existing natural gas transmission pipelines and establish potential pipeline modifications for a natural gas transmission pipeline network to meet ASME B31.12 design requirements, BlendPATH does NOT qualify the examined pipeline network nor modifications thereof for hydrogen blending. Further actual natural gas transmission pipeline network assessment and qualification will require additional evaluation, independent of BlendPATH application, as specified in ASME B31.12. It is the sole responsibility of the pipeline owner and operator to ensure that their pipeline qualifies for the ASME B31.12 design option that they choose and any other relevant regulations and that any defects present in the pipeline are acceptable at the operating pressure chosen.\n----------------------------\n'

@dataclass
class BlendPATH_scenario():
    '''
    A class to create a BlendPATH scenario

    Attributes:
    -----------
    casestudy_name: str
        Case study name - must also exist in case study folder
    save_directory : str
        Save directory for output files
    design_option : str
        Design option for initial assessment - ['No fracture control','A','B']
    location_class : int
        Location class for ASME B31.12
    joint_factor : int
        Joint factor for ASME B31.12
    design_pressure_known : bool
        Design pressure known?
    pressure_drop_factor : float
        Scalar adjustment to pressure drop calculation
    compression_ratio : float
        Compressor compression ratio
    nodes : int
        Number of nodes for discretization in momentum equation
    blending : float
        Fraction of h2 in h2/natural gas mixture
    natural_gas_cost : float
        $/MMBTU
    hydrogen_cost : float
        $/kg
    verbose : bool
        Verbose outputs (future use)
    final_outlet_pressure : int
        Pressure at pipeline exit (Pa)
    '''
    casestudy_name : str = 'Wangetal2018'           # Case study name - must also exist in case study folder
    save_directory : str = 'out/'                   # Save directory for output files
    design_option : str = 'no fracture control'     # Design option for initial assessment
    location_class : int = 1                        # Location class for ASME B31.12
    joint_factor : int = 1                          # Joint factor for ASME B31.12
    design_pressure_known : bool = False            # Design pressure known?
    pressure_drop_factor : float = 1                # Scalar adjustment to pressure drop calculation
    compression_ratio : float = 1.55                # Compressor compression ratio
    nodes : int = 50                                # Number of nodes for discretization in momentum equation
    blending : float = 0.20                         # Fraction of h2 in h2/ch4 mixture
    natural_gas_cost : float = 4.00                 # $/MMBTU
    hydrogen_cost : float = 1.00                    # $/kg
    verbose : bool = True
    final_outlet_pressure : int = 2000000           # Pa
    region : str = 'GP'                             # Region for ANL misc, ROW, and labor costs
    batch_id : int = 0
    PL_num_diams : int = 5                          # Number of diameters to check for the parallel loop method
    original_CS_rating : dict = field(default_factory=lambda:{'CS_1':12.5,'CS_2':12.5,'CS_3':12.5}) # Original rating of compressors
    inline_inspection_interval: float = 3           # Interval in years for conducting inline-inspections

    def __post_init__(self):

        if self.verbose:
            print(disclaimer_message())

        self.casestudy_path = f'{self.casestudy_name}/'
        self.path = self.casestudy_path+self.casestudy_name
        self.density_steel = 7840
        self.steel_costs_per_kg = pd.read_csv(resource_filename(__name__,'resources/steel_costs_per_kg.csv'),index_col = None,header = 1)
        self.params = {'design_option':self.design_option,
                'location_class':self.location_class,
                'joint_factor':self.joint_factor,
                'density_steel':self.density_steel,
                'pressure_drop_factor':self.pressure_drop_factor,
                'compression_ratio':self.compression_ratio,
                'nodes':self.nodes,
                'PL_num_diams':self.PL_num_diams}
        self.design_option_original = self.design_option

        #   Calculate capacity of pipeline based on sum of offtakes
        gsce_out_df=pd.read_excel(f'{self.casestudy_path}/{self.casestudy_name}_Event.xlsx','GSCE')
        vals = gsce_out_df[['Parameter','Value']]
        capacity = 0
        for i,x in enumerate(vals['Parameter']):
            if (('GDEM' in x) and ('QSET' in x)):
                capacity += vals['Value'][i]
        capacity = capacity*3.412141633*24 # convert to MMBTU-day
        self.gsce_out_df = vals
        self.capacity = capacity


        # Create a new directory for saving if it doesn't exist yet
        for filepaths in [self.save_directory,f'{self.save_directory}NetworkFiles',f'{self.save_directory}ResultFiles']:
            if not os.path.exists(filepaths):
                os.makedirs(filepaths)

        # Copy files to save directory
        base_name =  f'{self.save_directory}NetworkFiles/{self.batch_id}_{self.casestudy_name}'
        new_file_name = f'{base_name}_Network_blending_in.xlsx'
        new_file_name_event = f'{base_name}_Event_blending_in.xlsx'
        shutil.copy(f'{self.casestudy_path}/{self.casestudy_name}_Network.xlsx',new_file_name)
        shutil.copy(f'{self.casestudy_path}/{self.casestudy_name}_Event.xlsx',new_file_name_event)
        
        #   Update blending amount/composition in network file
        GCUS = pd.read_excel(new_file_name,'GCUS')
        if not (GCUS['[QUAL]']=='H2').any():
            GCUS = pd.concat((GCUS,pd.DataFrame({'[QUAL]':'H2','H00':0},index=[0])))
        gsup_val = f'H{int(100*self.blending):02d}' # Sets to 2 digits and adds any leading zeros
        GCUS[gsup_val] = GCUS.apply(lambda x: x['H00']*(1-self.blending) if x['[QUAL]'] != 'H2' else self.blending*100 ,axis=1)
        GCUS = GCUS[['[QUAL]',gsup_val]]
        
        #   Update blending amount in GSUP sheet
        GSUP = pd.read_excel(new_file_name,'GSUP')
        GSUP['SQSETNAME  = DEFAULT'] = gsup_val

        # Write to excel file
        workbook1 = openpyxl.load_workbook(new_file_name)
        del workbook1['GSUP']
        workbook1.save(new_file_name)
        with pd.ExcelWriter(new_file_name,engine='openpyxl',mode="a",if_sheet_exists='replace') as writer:  
            GSUP.to_excel(writer, sheet_name='GSUP',index=False)
            GCUS.to_excel(writer, sheet_name='GCUS',index=False)


        # Run initial SAInt simulation
        self.dllfolder = "C:/Program Files/encoord/SAInt-v3/"
        mydll = cdll.LoadLibrary(self.dllfolder + "SAInt-API.dll")
        mydll.evalStr.restype=c_wchar_p
        mydll.evalInt.restype=c_int
        mydll.evalCmdInt.restype=c_int
        mydll.evalBool.restype=c_bool
        mydll.evalFloat.restype=c_float
        
        daf.runSAInt(mydll,base_name,'_blending_in')
        
        #   Get critical properties
        self.params['T_c'] = mydll.evalFloat('GNO.0.NQ!TC.[K]')
        self.params['P_c'] = mydll.evalFloat('GNO.0.NQ!PC.[Pa]')
        self.params['viscosity'] = mydll.evalFloat('GNET.VISC')
        gcs_df=pd.read_excel(new_file_name,'GCS')
        self.init_CS_names = gcs_df['Name'].values
        self.init_CS_fuel,_,_ = daf.get_compressor_usage(mydll,self.init_CS_names,'kg/s') # kg/s for fuel usage on initial run
        self.GCV_fuel = mydll.evalFloat('GNO.0.NQ!GCV')

        self.original_CS_rating = list({k:self.original_CS_rating[k] for k in self.init_CS_names}.values()) # Save the original compressor ratings. Needed to see if upgrades are required

        
        #
        # PIPE ASSESSMENT
        #
        #   Get the number of pipes
        self.numPipes = mydll.evalCmdInt("ecount('GPI.%')")

        #   Get pipe params and joing with mechanical properties
        self.pipe_params = daf.getpipeParams(mydll,self.numPipes)
        mechanical_props = pd.read_csv(f'{self.path}_pipe_mech_props.csv',index_col = None,header = 0)
        self.pipe_params = self.pipe_params.join(mechanical_props.set_index('Pipe Name'),on = 'Pipe Name')

        # Get node fluid props
        self.node_fluidprops = daf.getnodefluidprops(mydll,self.numPipes)

        # Get segment props
        self.design_pressure_ASME = daf.get_design_pressure_ASME(self.numPipes,self.design_option,self.location_class,self.joint_factor,self.pipe_params)

        self.segment_props = daf.segment_props(mydll,self.casestudy_path,self.casestudy_name,self.design_pressure_ASME,self.pipe_params,self.numPipes,self.compression_ratio,self.final_outlet_pressure)


        #----------- Identify pipe segments that should have the same design pressure throughout ---------------------
        self.pipe_segments = daf.pipesegments(self.casestudy_path,self.casestudy_name)
        # -------------------- Assess Design Pressure Violations ----------------------------
        self.design_pressure_violations = daf.get_design_pressure_violations(self.pipe_params,self.design_pressure_ASME,self.pipe_segments,self.params,self.segment_props)
        #--------------------- Yield Strengths and Schedules ---------------------------
        self.yield_strengths,self.schedules = daf.get_yields_and_schedules()
        #------------------------- Assign design pressure to pipes ----------------------------
        self.design_pressure = daf.get_design_pressure(self.design_pressure_known,self.pipe_segments,self.path,self.pipe_params)
        #--------------- Identify demand nodes in pipe segments ----------------------------
        self.network_pipes,self.segment_demand_nodes = daf.get_demand_nodes(self.pipe_segments,self.path)

        self.pipe_params = self.pipe_params.join(self.design_pressure_ASME)
        self.max_node_pressure = daf.setNodeMaxPressure(mydll,self.pipe_params,self.numPipes)
        self.max_node_pressure_original = self.max_node_pressure

        #   Add grade to pipe params
        self.pipe_params = self.pipe_params.merge(self.yield_strengths[['Grade','SMYS [Mpa]']].rename(columns={'SMYS [Mpa]':'Yield strength [Mpa]'}), how = 'inner', on='Yield strength [Mpa]')
        
        # Add schedule to pipe params
        relevant_schedules = self.schedules.loc[self.schedules['DN'].isin(self.pipe_params['DN'])] 
        sch_list = []
        for pipe_index in self.pipe_params.index:
            DN = self.pipe_params["DN"][pipe_index]
            thickness = self.pipe_params["Wall thickness [mm]"][pipe_index]
            pipe_schedules = relevant_schedules.loc[relevant_schedules['DN'] == DN].where(relevant_schedules == round(thickness,3)).dropna(how='all',axis=1).columns
            if pipe_schedules.empty:
                pipe_min_schedule = "Custom - Thickness does not correspond to a schedule"
            else:
                pipe_min_schedule = pipe_schedules[0]
            sch_list.append(pipe_min_schedule)

        self.pipe_params = self.pipe_params.join(pd.DataFrame(sch_list, columns= ["Schedule"]))

        #   Save access to dll for future use
        self.mydll_original = mydll

    def __repr__(self):
        return f'{self.casestudy_name} BlendPATH scenario \n\t Design option: {self.design_option}'
    
    def run_mod(self,mod_type,design_option='b'):
        mod_type_filter = mod_type.lower().replace(' ','_')
        if mod_type_filter not in ['direct_replacement','parallel_loop','additional_compressors']:
            print(f'{mod_type} is an invalid modification. Try direct_replacement, parallel_loop or additional_compressors')
            return
        self.mod_type = mod_type_filter
        self.design_option = design_option
        self.params['design_option'] = design_option
        
        error = np.inf
        loops = 0
        original_mass_flowrate = {'seg_in':self.segment_props['Outlet Mass flow rate [kg/s]'],\
                                  'seg_out':self.segment_props['Inlet Mass flow rate [kg/s]']}
        
        self.max_node_pressure = self.max_node_pressure_original
         
        
        #----------------- Direct replacement module:  ----------------- 
        if mod_type_filter=='direct_replacement':

                    
        
            pipe_material_cost_direct_replacement,schedules_minviable,thicknesses,new_pressures = \
                mod_fxn.direct_replacement(self.yield_strengths,self.pipe_params,self.design_pressure,self.schedules,\
                                        self.design_pressure_violations,self.steel_costs_per_kg,self.params,self.segment_props)

            min_costs,assoc_grade,min_DNs,_ = self.get_min_cost(mod_type_filter,pipe_material_cost_direct_replacement,schedules_minviable)

            # geom_out = {'DN':[],'lengths':[],'wth':[],'inner diameter':[],'outer diameter':[],'material cost':min_costs,'Grade':assoc_grade,'Schedule':min_DNs,'MAOP':[]}
            geom_out = {key:[] for key in ['DN','lengths','wth','inner diameter','outer diameter','material cost','Grade','Schedule','MAOP']}
            for i in self.pipe_params['Pipe Name'].values:
                
                pipe_params_row = self.pipe_params.loc[self.pipe_params['Pipe Name']==i,:]
                geom_out['DN'].append(pipe_params_row['DN'].iat[0])
                geom_out['lengths'].append(pipe_params_row['Length [km]'].iat[0]*1000)
                # geom_out['inner diameter'].append(pipe_params_row['Pipe Diameter [mm]'].iat[0])
                geom_out['outer diameter'].append(pipe_params_row['Outer Diameter [mm]'].iat[0])
                if i in assoc_grade:
                    min_grade = assoc_grade[i]
                    geom_out['wth'].append(thicknesses.loc[i,min_grade]*1000)
                    geom_out['material cost'].append(min_costs[i])
                    geom_out['Grade'].append(assoc_grade[i])
                    geom_out['Schedule'].append(min_DNs[i])
                    geom_out['MAOP'].append(new_pressures.loc[i,min_grade])
                else:
                    geom_out['wth'].append(pipe_params_row['Wall thickness [mm]'].iat[0])
                    geom_out['material cost'].append(0)
                    geom_out['Grade'].append('')
                    geom_out['Schedule'].append('')
                    geom_out['MAOP'].append(pipe_params_row['ASME Design Pressure [MPa]'].iat[0])
                geom_out['inner diameter'].append(geom_out['outer diameter'][-1]-2*geom_out['wth'][-1])

            pipe_names = self.pipe_params['Pipe Name'].values
            MAOP_df = pd.DataFrame({'Pipe Name':pipe_names,'ASME Design Pressure [MPa]':geom_out['MAOP']})
            self.max_node_pressure = daf.setNodeMaxPressure(self.mydll_original,MAOP_df,self.numPipes)

        #----------------------- Parallel loop module -----------------------------------------
        elif mod_type_filter=='parallel_loop':
            
            while error>0.01:
                p_out = -1
                m_dot = -1
                cs_fuel_use = []
                for index,row in self.segment_props.iterrows():
                    p_prev = p_out
                    m_prev = m_dot
                    p_in = row['Inlet pressure [Pa]']
                    p_out = row['Outlet pressure [Pa]']
                    m_dot = row['Outlet Mass flow rate [kg/s]']
                    if index==0:
                        continue
                    X = f"H2:{self.blending},CH4:{1-self.blending}"
                    T = 15+273.15
                    cs_fuel_use.append(daf.get_fuel_flow(p_prev,p_in,X,T,self.GCV_fuel,m_prev))
                cs_fuel_use.append(0) # Last segment does not have fuel extration downstream of it
                self.segment_props['Outlet Mass flow rate [kg/s]'] = original_mass_flowrate['seg_in'] + cs_fuel_use
                self.segment_props['Inlet Mass flow rate [kg/s]'] = original_mass_flowrate['seg_out'] + cs_fuel_use
                self.node_fluidprops['Mass flow rate [kg/s]'].iat[0] = self.segment_props['Inlet Mass flow rate [kg/s]'].iat[0]

            
                length_loop_by_segment,schedule_by_segment,pipe_material_cost_PL_by_segment,dn_options_all_segments,thicknesses,inner_diams = \
                    mod_fxn.parallel_loop(self.node_fluidprops,self.segment_props,self.path,self.pipe_segments,self.yield_strengths,self.schedules,\
                                        self.segment_demand_nodes,self.network_pipes,self.pipe_params,self.steel_costs_per_kg,self.params)                 

                error  = (abs((sum(self.init_CS_fuel)-sum(cs_fuel_use))/sum(self.init_CS_fuel))) if len(self.init_CS_names)>0 else 0
                loops += 1
                if loops>10:
                    break
                self.init_CS_fuel = cs_fuel_use

            min_costs,assoc_grade,min_DNs,min_index = self.get_min_cost(mod_type_filter,pd.DataFrame(pipe_material_cost_PL_by_segment))

            geom_out = {'DN':[],'lengths':[],'wth':[],'inner diameter':[],'material cost':min_costs,'Grade':assoc_grade,'Schedule':[]}
            for i in range(len(self.segment_props.index)):
                # If statements are added below to handle cases where no looping is needed in a given segment.
                grade_index = -1 if len(assoc_grade)<1 else self.yield_strengths.loc[self.yield_strengths['Grade']==assoc_grade[i]].index.tolist()[0]
                if min_costs[i] == np.inf:
                    min_costs[i] = 0

                geom_out['DN'].append(dn_options_all_segments[i][min_DNs[i]])
                if (pd.isna(length_loop_by_segment[i][min_DNs[i]]['Loop length [m]'].iat[min_index[i]]) or length_loop_by_segment[i][min_DNs[i]].empty):
                    geom_out['lengths'].append(0)
                else:
                    geom_out['lengths'].append(length_loop_by_segment[i][min_DNs[i]]['Loop length [m]'].iat[min_index[i]])
                geom_out['wth'].append(0 if grade_index == -1 else thicknesses[i][min_DNs[i]][grade_index])
                geom_out['inner diameter'].append(0 if grade_index == -1 else inner_diams[i][min_DNs[i]][grade_index])
                geom_out['Schedule'].append(schedule_by_segment[i][min_DNs[i]]['Schedules'].iat[min_index[i]])
            
        #----------------------- Additional compressors module --------------------------
        elif mod_type_filter=='additional_compressors':
        
            X = f"H2:{self.blending},CH4:{1-self.blending}"
            T = 15+273.15

            loops = 0
            while error>0.01:

                num_compressors_by_segment,compressor_length_by_segment,compression_ratio_by_segment,pressures = \
                    mod_fxn.additional_compressors(self.node_fluidprops,self.segment_props,self.pipe_segments,self.segment_demand_nodes,self.network_pipes,self.pipe_params,self.params)

                cs_fuel_use = []
                for index,row in self.segment_props.iterrows():
                    cs_fuel_seg = 0
                    for p in pressures[index]:
                        p_in,p_out = p
                        m_dot = row['Outlet Mass flow rate [kg/s]']
                        cs_fuel_seg += daf.get_fuel_flow(p_out,p_in,X,T,self.GCV_fuel,m_dot)
                    cs_fuel_use.append(cs_fuel_seg)
                self.segment_props['Inlet Mass flow rate [kg/s]'] = original_mass_flowrate['seg_out'] + cs_fuel_use
                self.node_fluidprops['Mass flow rate [kg/s]'].iat[0] = self.segment_props['Inlet Mass flow rate [kg/s]'].iat[0]
                
                error  = (abs((sum(self.init_CS_fuel)-sum(cs_fuel_use))/sum(self.init_CS_fuel))) if (len(self.init_CS_names)>0 and sum(self.init_CS_fuel)>0) else 0
                self.init_CS_fuel = cs_fuel_use
                loops += 1
                if loops > 10:
                    break
            
            mat_cost = 0 #Not relevant for this method
            geom=0

            geom_out = {'compression_ratio':compression_ratio_by_segment,'compressor_length':compressor_length_by_segment}

        #   Post processing for geometry (ignore for additional compressors)
        if not mod_type_filter=='additional_compressors':
            geom = pd.DataFrame(geom_out)
            geom = geom.loc[geom['material cost']>0]
            geom_grades = geom.groupby('DN').agg({'Grade':'first'})
            geom_schedules = geom.groupby('DN').agg({'Schedule':'first'})
            geom = pd.DataFrame({'Length [mi]':geom.groupby('DN')['lengths'].sum()/1609.34,\
                                'Material Cost [$]':geom.groupby('DN')['material cost'].sum()}).reset_index()
            geom = pd.merge(geom,geom_grades,on='DN')
            geom = pd.merge(geom,geom_schedules,on='DN')
            geom['Pipe Diameter [in]'] = np.ceil(geom['DN'] * 0.0393701)
            geom['Material unit cost [$/in-mile]'] = geom['Material Cost [$]']/geom['Pipe Diameter [in]']/geom['Length [mi]']
            mat_cost = sum(min_costs)

            
            


        detailed_geom = pd.DataFrame(geom_out)

        #   Remake the network file and run it
        #   Prepare mod_out for the finanacial analysis
        self.mod_out = {'geom':geom,'detailed_geom':detailed_geom,'mat_cost':mat_cost}
        CS_names = self.remake_network_file(geom_out,mod_type_filter)

        #   Rerun SAInt using modified file and update output dictionary
        self.mod_out.update(self.runModifiedFile(mod_type_filter,CS_names))
        

    def get_min_cost(self,mod_type,df_in,schedules_minviable=False):
        
        if df_in.empty:
            min_costs = []
            assoc_grade = []
            for j in range(len(self.pipe_segments)):
                min_costs.append(0)
                assoc_grade.append(None)
            return min_costs,assoc_grade,[],[]
        
        if mod_type=='direct_replacement':
            min_cost_df = df_in.copy().fillna(np.inf).astype('float64')
            min_cost_df['Min Cost [$]'] = min_cost_df.min(axis=1)
            min_cost_df['Associated Grade'] = min_cost_df.idxmin(axis=1)
            min_cost_df['Associated Sch'] = min_cost_df.apply(lambda x: schedules_minviable.loc[x.name,x['Associated Grade']],axis=1)

            min_costs =  min_cost_df['Min Cost [$]']
            assoc_grade = min_cost_df['Associated Grade']
            min_index = min_cost_df['Associated Sch']
            min_index_2 = -1


        elif mod_type=='parallel_loop':
            
            min_costs = []
            assoc_grade = []
            min_index = []
            min_index_2 = []
            for index,row in df_in.iterrows(): # Loop thru pipe segments
                min_cost_per_segment = np.inf
                assoc_grade_per_segment = 'X42'
                min_DN_per_segment = -1
                min_index_per_item = -1
                for index,item in row.items(): # Loop thru diameter
                    a = item['Pipe material costs [$]'].min(skipna=True)
                    if a==np.NaN:
                        continue
                    if a < min_cost_per_segment:
                        min_index_per_item = item['Pipe material costs [$]'].idxmin(skipna=True)
                        min_cost_per_segment = a
                        assoc_grade_per_segment = item['Grade'].iat[min_index_per_item]
                        min_DN_per_segment = index
                        min_index_per_item = item['Pipe material costs [$]'].idxmin(skipna=True)

                min_costs.append(min_cost_per_segment)
                assoc_grade.append(assoc_grade_per_segment)
                min_index.append(min_DN_per_segment)
                min_index_2.append(min_index_per_item)
                    
        return min_costs,assoc_grade,min_index,min_index_2

    def run_financial(self,cost_overrides=None,user_params=None):
        cap_costing = self.mod_out['geom']

        if cost_overrides == None:
            override_labor = np.NaN
            override_row = np.NaN
            override_misc = np.NaN
            override_orig_pipeline = np.NaN
        else:
            override_labor = cost_overrides['OVERRIDE: labor cost ($/in/mi)']
            override_row = cost_overrides['OVERRIDE: row cost ($/in/mi)']
            override_misc = cost_overrides['OVERRIDE: misc cost ($/in/mi)']
            override_orig_pipeline = cost_overrides['OVERRIDE: Original pipeline cost ($)']
    
        #   ANL costing coefficients for all regions. Material cost is ignored as internal methods are used
        anl_coefs_regional = {'GP':{'labor':[10406,0.20953,-0.08419],'misc':[4944,0.17351,-0.07621],'ROW':[2751,-0.28294,0.00731],'Material':[5813,0.31599,-0.00376]},
                              'NE':{'labor':[249131,-0.33162,-0.17892],'misc':[65990,-0.29673,-0.06856],'ROW':[83124,-0.66357,-0.07544],'Material':[10409,0.296847,-0.07257]},
                              'MA':{'labor':[43692,0.05683,-0.10108],'misc':[14616,0.16354,-0.16186],'ROW':[1942,0.17394,-0.01555],'Material':[9113,0.279875,-0.00840]},
                              'GL':{'labor':[58154,-0.14821,-0.10596],'misc':[41238,-0.34751,-0.11104],'ROW':[14259,-0.65318,0.06865],'Material':[8971,0.255012,-0.03138]},
                              'RM':{'labor':[10406,0.20953,-0.08419],'misc':[4944,0.17351,-0.07621],'ROW':[2751,-0.28294,0.00731],'Material':[5813,0.31599,-0.00376]},
                              'SE':{'labor':[32094,0.06110,-0.14828],'misc':[11270,0.19077,-0.13669],'ROW':[9531,-0.37284,0.02616],'Material':[6207,0.38224,-0.05211]},
                              'PN':{'labor':[32094,0.06110,-0.14828],'misc':[11270,0.19077,-0.13669],'ROW':[9531,-0.37284,0.02616],'Material':[6207,0.38224,-0.05211]},
                              'SW':{'labor':[95295,-0.53848,0.03070],'misc':[19211,-0.14178,-0.04697],'ROW':[72634,-1.07566,0.05284],'Material':[5605,0.41642,-0.06441]},
                              'CA':{'labor':[95295,-0.53848,0.03070],'misc':[19211,-0.14178,-0.04697],'ROW':[72634,-1.07566,0.05284],'Material':[5605,0.41642,-0.06441]}}
        
        #   Set ProFAST financial parameters. Default if none are entered
        financial_params_default = {'analysis_start_year':2020,'operating_life':50,'installation_months':36,'long_term_utilization':1,'property_tax_and_insurance':0.009,
                            'admin_expense':0.005,'total_income_tax_rate':0.2574,'capital_gains_tax_rate':0.15,'general_inflation_rate':0.019,'leverage_after_tax_nominal_discount_rate':0.1,
                            'debt_equity_ratio_of_initial_financing':1.5,'debt_type':'Revolving debt','loan_period_if_used':0,'debt_interest_rate':0.037,'cash_onhand':1}
        financial_params = financial_params_default if user_params==None else user_params

        anl_coefs = anl_coefs_regional[self.region]
        #   ANL costing is in 2018 dollars
        year_adj = 596.2/603.1 # Adj from 2018 to 2020 $

        def get_compressor_cost(avg_station_cap,num_comp_stations,revamped_ratio=1):
            comp_coef = {'Material':[3175286,532.7853,0.0010416],'Labor':[1581740,299.2887,0.001142],'Misc':[1696686,184.1443,0.0018417],'Land':[66216.72,0,0.0001799]}
            comp_cost_i = []
            for i in range(len(avg_station_cap)):
                station_cap_array = [1,avg_station_cap[i],avg_station_cap[i]**2]
                material = np.dot(comp_coef['Material'],station_cap_array)
                labor = np.dot(comp_coef['Labor'],station_cap_array)
                misc =  np.dot(comp_coef['Misc'],station_cap_array)
                land = np.dot(comp_coef['Land'],station_cap_array) if revamped_ratio==1 else 0
                comp_cost_i.append((material+labor+misc+land)*596.2/575.4*revamped_ratio) # Go from 2008 to 2020
            comp_cost_i = np.multiply(comp_cost_i,num_comp_stations)
            comp_cost = sum(comp_cost_i) 
            return comp_cost_i

        def get_original_pipe_cost():
            '''
                This calculates the pipe cost of the pre existing pipeline
            '''
            orig_pipe_params = self.pipe_params.copy()
            orig_pipe_params['ID [m]'] = orig_pipe_params['Pipe Diameter [mm]']/1000
            orig_pipe_params['th [m]'] = orig_pipe_params['Wall thickness [mm]']/1000
            orig_pipe_params['OD [m]'] = orig_pipe_params['ID [m]']+2*orig_pipe_params['th [m]']
            orig_pipe_params['length [m]'] = orig_pipe_params['Length [km]']*1000
            orig_pipe_params['volume [m3]'] =  np.pi*(orig_pipe_params['OD [m]']**2- orig_pipe_params['ID [m]']**2)/4*orig_pipe_params['length [m]']
            orig_pipe_params['weight [kg]'] = orig_pipe_params['volume [m3]']*self.params['density_steel']

            orig_pipe_params = pd.merge(orig_pipe_params,self.steel_costs_per_kg[['Grade','Price [$/kg]']],on='Grade')

            orig_pipe_params['Material Cost [$]'] = orig_pipe_params['Price [$/kg]']*orig_pipe_params['weight [kg]']

            orig_cap_costing = pd.DataFrame({'Length [mi]':orig_pipe_params.groupby(['Grade','Schedule','OD [m]'])['length [m]'].sum()/1609.34,\
                                'Material Cost [$]':orig_pipe_params.groupby(['Grade','Schedule','OD [m]'])['Material Cost [$]'].sum()}).reset_index()
            orig_cap_costing['Pipe Diameter [in]']=np.round(orig_cap_costing['OD [m]']*39.3701)

            orig_cap_costing['Material unit cost [$/in-mile]'] = orig_cap_costing['Material Cost [$]']/orig_cap_costing['Pipe Diameter [in]']/orig_cap_costing['Length [mi]']
            orig_cap_costing['Labor unit cost [$/in-mile]'] = orig_cap_costing.apply(lambda x: anl_coefs['labor'][0]*(x['Pipe Diameter [in]']**anl_coefs['labor'][1])*(x['Length [mi]']**anl_coefs['labor'][2]) ,axis=1)*year_adj
            orig_cap_costing['Misc unit cost [$/in-mile]'] = orig_cap_costing.apply(lambda x: anl_coefs['misc'][0]*(x['Pipe Diameter [in]']**anl_coefs['misc'][1])*(x['Length [mi]']**anl_coefs['misc'][2]) ,axis=1)*year_adj
            orig_cap_costing['ROW unit cost [$/in-mile]'] = orig_cap_costing.apply(lambda x: anl_coefs['ROW'][0]*(x['Pipe Diameter [in]']**anl_coefs['ROW'][1])*(x['Length [mi]']**anl_coefs['ROW'][2]) ,axis=1)*year_adj
            orig_cap_costing['Labor [$]'] = orig_cap_costing['Labor unit cost [$/in-mile]'] * orig_cap_costing['Pipe Diameter [in]'] * orig_cap_costing['Length [mi]']
            orig_cap_costing['Misc [$]'] = orig_cap_costing['Misc unit cost [$/in-mile]'] * orig_cap_costing['Pipe Diameter [in]'] * orig_cap_costing['Length [mi]']
            orig_cap_costing['ROW [$]'] = orig_cap_costing['ROW unit cost [$/in-mile]'] * orig_cap_costing['Pipe Diameter [in]'] * orig_cap_costing['Length [mi]']

            total_cap =  orig_cap_costing['Material Cost [$]'].sum() + orig_cap_costing['Labor [$]'].sum() + orig_cap_costing['Misc [$]'].sum() + orig_cap_costing['ROW [$]'].sum()
            return total_cap,orig_cap_costing
        
        #   If pipe is added then calculate pipe costs using ANL coefficient for GP region (Brown 2022)
        small_positive = 1e-7 # This allows solution and file output writing when length of a given DN is set to zero, but usually this is a sign of an issue somewhere
        if not self.mod_type=='additional_compressors':
            cap_costing['Labor unit cost [$/in-mile]'] = override_labor if not pd.isna(override_labor) else cap_costing.apply(lambda x: anl_coefs['labor'][0]*((x['Pipe Diameter [in]']+small_positive)**anl_coefs['labor'][1])*((x['Length [mi]']+small_positive)**anl_coefs['labor'][2]),axis=1)*year_adj
            cap_costing['Misc unit cost [$/in-mile]'] = override_misc if not pd.isna(override_misc) else cap_costing.apply(lambda x: anl_coefs['misc'][0]*((x['Pipe Diameter [in]']+small_positive)**anl_coefs['misc'][1])*((x['Length [mi]']+small_positive)**anl_coefs['misc'][2]) ,axis=1)*year_adj
            cap_costing['ROW unit cost [$/in-mile]'] = override_row if not pd.isna(override_row) else cap_costing.apply(lambda x: anl_coefs['ROW'][0]*((x['Pipe Diameter [in]']+small_positive)**anl_coefs['ROW'][1])*((x['Length [mi]']+small_positive)**anl_coefs['ROW'][2]) ,axis=1)*year_adj
            cap_costing['Labor [$]'] = cap_costing['Labor unit cost [$/in-mile]'] * cap_costing['Pipe Diameter [in]'] * cap_costing['Length [mi]']
            cap_costing['Misc [$]'] = cap_costing['Misc unit cost [$/in-mile]'] * cap_costing['Pipe Diameter [in]'] * cap_costing['Length [mi]']
            cap_costing['ROW [$]'] = cap_costing['ROW unit cost [$/in-mile]'] * cap_costing['Pipe Diameter [in]'] * cap_costing['Length [mi]']

            non_mat_costs = cap_costing['Labor [$]'].sum() + cap_costing['Misc [$]'].sum()
            if self.mod_type=='parallel_loop':
                non_mat_costs = non_mat_costs+cap_costing['ROW [$]'].sum()
            else:
                cap_costing[['ROW [$]','ROW unit cost [$/in-mile]']] = 0
            total_cap = self.mod_out['mat_cost'] + non_mat_costs
            pipeline_added = cap_costing['Length [mi]'].sum()

            cap_costing_out = cap_costing[['Pipe Diameter [in]','Length [mi]','Grade','Schedule','Material Cost [$]','Labor [$]','Misc [$]','ROW [$]',\
                            'Material unit cost [$/in-mile]','Labor unit cost [$/in-mile]','Misc unit cost [$/in-mile]','ROW unit cost [$/in-mile]']]

        else:  
            total_cap = 0
            pipeline_added = 0

        MW2hp = 1341.02
        hp_max,hp_min = 30000,3000

        #   Calculate original pipeline and compressor costs
        original_CS_rating_hp = np.multiply(self.original_CS_rating,MW2hp)
        num_comp_stations = 1+np.floor(original_CS_rating_hp/hp_max)
        avg_station_cap = [hp_min if x<hp_min else x/num_comp_stations[i] for i,x in enumerate(original_CS_rating_hp)]

        original_cap_pipes,orig_cap_costing = get_original_pipe_cost()

        if self.mod_type=='direct_replacement' and self.blending<=0: # self.blending limit was previously set to 10% to account for potential limits w/ compressor speed. 
                                                                     # However, fluid properties (vol. flow rate, low inlet pressure) for blending are assumed to change  
                                                                     # significantly such that you may need to refurb compressor anyway 
            revamped_compressors_i = get_compressor_cost(avg_station_cap,num_comp_stations,revamped_ratio=0.16)
        else:
            revamped_compressors_i = get_compressor_cost(avg_station_cap,num_comp_stations,revamped_ratio=0.66)
        revamped_compressors = sum(revamped_compressors_i)

        #   Calculate cost of compressors after modification
        num_new_compressors = len(self.mod_out['CS_rating'])-len(self.original_CS_rating)
        padded_original_ratings = np.pad(self.original_CS_rating,(0,num_new_compressors),'constant')
        added_CS_rating = np.maximum(np.multiply(self.mod_out['CS_rating']-padded_original_ratings,MW2hp),0)
        num_comp_stations_new = (1+np.floor(added_CS_rating/hp_max)) *(added_CS_rating>0)
        avg_station_cap_new = [hp_min if x<hp_min else x/num_comp_stations_new[i] for i,x in enumerate(added_CS_rating)] * (added_CS_rating>0)
        added_cap_compressors_i = get_compressor_cost(avg_station_cap_new,num_comp_stations_new)
        added_cap_compressors = sum(added_cap_compressors_i)

        #   Compressor fuel usage (this is divided by the sum off-takes for capacity)
        compressor_fuel = sum(self.mod_out['CS_fuel'])/(self.capacity/24) # MMBTU/MMBTU
        
        #   Determine cost of compressor fuel (h2 cost is tabulated from external spreadsheets)
        H2_energy = 141.88
        nat_gas_cost = self.natural_gas_cost # $/MMBTU
        H2_cost = self.hydrogen_cost #$/kg

        GCV_NG_MJpsm3 = 41.66196841
        GCV_H2_MJpsm3 = 12.761
        blend_ratio_energy = (self.blending*GCV_H2_MJpsm3)/((self.blending*GCV_H2_MJpsm3)+(1-self.blending)*GCV_NG_MJpsm3)
        cs_cost = nat_gas_cost*(1-blend_ratio_energy)+H2_cost/H2_energy*1055.05853* blend_ratio_energy # $/MMBTU

        #   Set up modified pipeline dimension data to base valve addition/replacement and inline inspesction costs on

        existing_pipe_dim = self.segment_props[['DN','Length [m]']]
        existing_pipe_dim.loc[:,'Length [m]'] = existing_pipe_dim.loc[:,'Length [m]']/1609.34
        existing_pipe_dim = existing_pipe_dim.rename(columns = {'Length [m]':'Length [mi]'})

        #       Account for pipeline dimension data associated to loop if using parallel loop method

        if self.mod_type == 'parallel_loop':
            added_pipe_dim = self.mod_out['detailed_geom'][['DN', 'lengths']]
            added_pipe_dim.loc[:,'lengths'] = added_pipe_dim.loc[:,'lengths']/1609.34
            added_pipe_dim = added_pipe_dim.rename(columns = {'lengths':'Length [mi]'})
        else: 
            added_pipe_dim = pd.DataFrame(None,None, columns = ['DN','Length [mi]'])

        #   Establish valve spacing interval as function of Class Location as per ASME B31.12 
        ASMEB3112_valve_spacing = {1: 20, 2: 15, 3: 10, 4: 5}       # Dict format with key being class location 

        #   Import valve costs as a df, apply valve costs for all methods assuming all block valves have to be replaced
        #   for existing pipe main and/or added for new pipe as in the case for direct replacement and parallel looping

        valve_costs_df = pd.read_csv(resource_filename(__name__,'resources/valve_costs.csv'),index_col = None,header = 1)

        existing_pipe_dim['Num Valves Replaced'] = (existing_pipe_dim['Length [mi]']/ASMEB3112_valve_spacing[self.location_class]).apply(np.floor)
        added_pipe_dim['Num Valves Added'] = (added_pipe_dim['Length [mi]']/ASMEB3112_valve_spacing[self.location_class]).apply(np.floor)

        pipe_main_valve_cost_df = pd.merge(existing_pipe_dim, valve_costs_df.loc[valve_costs_df['DN'].isin(existing_pipe_dim['DN']) & (valve_costs_df['Install Type'] == "Buried")], 'left', 'DN')
        pipe_main_valve_cost = (pipe_main_valve_cost_df['Num Valves Replaced'] * pipe_main_valve_cost_df['Installed Valve Cost [2020$]']).sum()

        pipe_loop_valve_cost = 0
        if self.mod_type == 'parallel_loop':
            pipe_loop_valve_cost_df = pd.merge(added_pipe_dim, valve_costs_df.loc[valve_costs_df['DN'].isin(existing_pipe_dim['DN']) & (valve_costs_df['Install Type'] == "Buried")], 'left', 'DN')
            pipe_loop_valve_cost = (pipe_loop_valve_cost_df['Num Valves Added'] * pipe_loop_valve_cost_df['Installed Valve Cost [2020$]']).sum() \

        #   Meter Station Modification Costing 
            
        #   Set up offtake data to base meter replacement costs on
        meters_df = self.gsce_out_df[['Parameter','Value']]
        
        #       Isolate pipeline offtakes for meter replacement
        for i,x in enumerate(meters_df['Parameter']):
            if not ((('GDEM' in x) and ('QSET' in x))):
                meters_df = meters_df.drop(i, axis=0)
        
        #       Import linear meter replacement cost parameters. Linear regression is in the for of 
        #       y = m*x + b     w/ b [=] 2020$ and m [=] 2020$/MMBTU-day
        Meter_replacement_cost_df = pd.read_csv(resource_filename(__name__,'resources/meter_replacement_cost_regression_parameters.csv'),index_col = None,header = 1)

        #       Cost meter replacements using linear cost regression on updated meter capacity data
        
        meters_df['Value'] = meters_df['Value']*3.412141633*24 # convert value from MW to MMBTU-day
        meters_df['Meter Replacement Cost'] = Meter_replacement_cost_df["m [2020$/MMBTU-day]"].iloc[0]*meters_df['Value']\
                                            + Meter_replacement_cost_df["b [2020$]"].iloc[0]

        meter_replacement_cost = meters_df['Meter Replacement Cost'].sum()

        #       Cost gas chromatography (GC) modifications (if present) or addition

        GC_cost_df = pd.read_csv(resource_filename(__name__,'resources/GC_cost.csv'),index_col = None,header = 1)
        Gas_chromatograph_cost = len(meters_df)*GC_cost_df['Installed Cost [2020$]'].iloc[0]

        #       Cost Pressure Regulator replacement
        regulator_cost_df = pd.read_csv(resource_filename(__name__,'resources/regulator_costs.csv'),index_col = None,header = 1)
        meters_df['Regulator Runs'] = (meters_df['Value']/regulator_cost_df['Capacity [MMBTU/day]'].iloc[0]).apply(np.ceil)
        meters_df['Regulator Replacement Cost'] = meters_df['Regulator Runs']*regulator_cost_df['Installed Regulator Cost [2020$]'].iloc[0]

        regulator_replacement_cost = meters_df['Regulator Replacement Cost'].sum()

        meter_regulator_station_modification_cost = meter_replacement_cost + Gas_chromatograph_cost + regulator_replacement_cost

        #   Import ILI costs per diameter as a dataframe
        ILI_cost_df = pd.read_csv(resource_filename(__name__,'resources/inline_inspection_costs.csv'),index_col = None,header = 1)

        #   Assess ILI costs for pipe mains in all methods, and pipe loops in parallel looping: 
        
        pipe_main_ILI_cost_df = pd.merge(existing_pipe_dim, ILI_cost_df.loc[ILI_cost_df['DN'].isin(existing_pipe_dim['DN'])], 'left', 'DN')
        pipe_main_ILI_cost = (pipe_main_ILI_cost_df['Length [mi]'] * pipe_main_ILI_cost_df['ILI cost [2020$/mi]']).sum()

        pipe_loop_ILI_cost = 0
        if self.mod_type == 'parallel_loop':
            pipe_loop_ILI_cost_df = pd.merge(added_pipe_dim, ILI_cost_df.loc[ILI_cost_df['DN'].isin(added_pipe_dim['DN'])], 'left', 'DN')
            pipe_loop_ILI_cost = (pipe_loop_ILI_cost_df['Length [mi]'] * pipe_loop_ILI_cost_df['ILI cost [2020$/mi]']).sum()

        #   Compile ILI costs and annualize for input to ProFAST
        network_ILI_cost = pipe_main_ILI_cost + pipe_loop_ILI_cost

        annualized_network_ILI_cost = network_ILI_cost/self.inline_inspection_interval
           
        #   For depreciation period value of new pipe and equipment, and undepreciated capital of existing equipment.
        depr_period = 30
        
        #   ProFAST setup
        ProFAST_json_file = f'{self.casestudy_path}{self.casestudy_name}.json'
        if not os.path.exists(ProFAST_json_file):
            pf = ProFAST.ProFAST()
            gen_inflation = financial_params['general_inflation_rate']
            pf.set_params('general inflation rate',gen_inflation)
            pf.set_params('analysis start year',financial_params['analysis_start_year'])
            pf.set_params('operating life',financial_params['operating_life'])
            pf.set_params('installation months',financial_params['installation_months'])
            pf.set_params('long term utilization',financial_params['long_term_utilization']) 
            pf.set_params('property tax and insurance',financial_params['property_tax_and_insurance'])
            pf.set_params('admin expense',financial_params['admin_expense'])
            pf.set_params('total income tax rate',financial_params['total_income_tax_rate'])
            pf.set_params('capital gains tax rate',financial_params['capital_gains_tax_rate'])
            pf.set_params('leverage after tax nominal discount rate',financial_params['leverage_after_tax_nominal_discount_rate'])
            pf.set_params('debt equity ratio of initial financing',financial_params['debt_equity_ratio_of_initial_financing']) 
            pf.set_params('debt type',financial_params['debt_type'])
            pf.set_params('loan period if used',financial_params['loan_period_if_used'])
            pf.set_params('debt interest rate',financial_params['debt_interest_rate']) 
            pf.set_params('cash onhand',financial_params['cash_onhand'])
        else:
            pf = ProFAST.ProFAST(ProFAST_json_file)
            gen_inflation = pf.vals['general inflation rate']
        pf.set_params('capacity',self.capacity) #units/day
        pf.set_params('commodity',{"name":'Natural Gas-H2 Blend',"unit":"MMBTU","initial price":0.25,"escalation":gen_inflation})
        pf.set_params('sell undepreciated cap',True)
        pf.set_params('tax losses monetized',True)
        
        #   Add compressor fuel feedstock
        pf.add_feedstock(name='Compressor Fuel',usage=compressor_fuel,unit='MMBTU',cost=cs_cost,escalation=gen_inflation)

        #   Add inline-inspection costs
        pf.add_fixed_cost(name='Inline inspection',usage=1,unit='$',cost=annualized_network_ILI_cost,escalation=gen_inflation)

        #   Set original network value
        original_network_residual_value = 0 
        if not pd.isna(override_orig_pipeline):
            original_network_residual_value = override_orig_pipeline

        if total_cap>0:
            if self.mod_type == 'parallel_loop':
                pf.add_capital_item(name='New pipe CAPEX',cost=total_cap,depr_type='Straight line',depr_period=depr_period,refurb=[0])
            elif self.mod_type == 'direct_replacement':
                pf.add_capital_item(name='New pipe CAPEX',cost=total_cap,depr_type='Straight line',depr_period=depr_period,refurb=[0])
        
        pf.add_capital_item(name='New compressor station CAPEX',cost=added_cap_compressors,depr_type='Straight line',depr_period=depr_period,refurb=[0])
        pf.add_capital_item(name='Compressor station refurbishment CAPEX',cost=revamped_compressors,depr_type='Straight line',depr_period=depr_period,refurb=[0])
        pf.add_capital_item(name='Original network residual value',cost=original_network_residual_value,depr_type='Straight line',depr_period=depr_period,refurb=[0])
        pf.add_capital_item(name='Meter & regulator station modification CAPEX',cost=meter_regulator_station_modification_cost,depr_type='Straight line',depr_period=depr_period,refurb=[0])
        pf.add_capital_item(name='Main-line valve replacement CAPEX',cost= pipe_main_valve_cost,depr_type='Straight line',depr_period=depr_period,refurb=[0])
        pf.add_capital_item(name='New loop-line valve CAPEX',cost=pipe_loop_valve_cost,depr_type='Straight line',depr_period=depr_period,refurb=[0])

        total_capex = total_cap + original_network_residual_value + added_cap_compressors + revamped_compressors + meter_regulator_station_modification_cost\
                    + pipe_main_valve_cost + pipe_loop_valve_cost
        
        if total_capex == 0:
            capex_fraction = {'New pipe':0,
                          'Original network':0,
                          'New compressor stations':0,
                          'Compressor station refurbishments':0,
                          'Meter & regulator station modifications': 0,
                          'Mainline valve replacements': 0,
                          'New loop line valves': 0}
        else:
            capex_fraction = {'New pipe':total_cap/total_capex,
                            'Original network':original_network_residual_value/total_capex,
                            'New compressor stations':added_cap_compressors/total_capex,
                            'Compressor station refurbishments':revamped_compressors/total_capex,
                            'Meter & regulator station modifications': meter_regulator_station_modification_cost/total_capex,
                            'Main-line valve replacements': pipe_main_valve_cost/total_capex,
                            'New loop-line valves': pipe_loop_valve_cost/total_capex}

        sol = pf.solve_price()
        cash_flow = pf.cash_flow(price = sol['lco'])
        price_breakdown = pf.get_cost_breakdown()

        # Calculate financial expense associated with equipment
        cap_expense = price_breakdown.loc[price_breakdown['Name']=='Repayment of debt','NPV'].tolist()[0]\
            + price_breakdown.loc[price_breakdown['Name']=='Interest expense','NPV'].tolist()[0]\
            + price_breakdown.loc[price_breakdown['Name']=='Dividends paid','NPV'].tolist()[0]\
            - price_breakdown.loc[price_breakdown['Name']=='Inflow of debt','NPV'].tolist()[0]\
            - price_breakdown.loc[price_breakdown['Name']=='Inflow of equity','NPV'].tolist()[0]  

        # Calculate remaining financial expenses
        remaining_financial = price_breakdown.loc[price_breakdown['Name']=='Non-depreciable assets','NPV'].tolist()[0]\
            + price_breakdown.loc[price_breakdown['Name']=='Cash on hand reserve','NPV'].tolist()[0]\
            - price_breakdown.loc[price_breakdown['Name']=='Sale of non-depreciable assets','NPV'].tolist()[0]\
            - price_breakdown.loc[price_breakdown['Name']=='Cash on hand recovery','NPV'].tolist()[0]

        if self.mod_type == 'additional_compressors':
            price_breakdown_newpipe = 0 
        else: 
            if total_cap==0:
                price_breakdown_newpipe = 0
            else:
                price_breakdown_newpipe = price_breakdown.loc[price_breakdown['Name']=='New pipe CAPEX','NPV'].tolist()[0]+cap_expense*capex_fraction['New pipe']
        price_breakdown_originalnetwork = price_breakdown.loc[price_breakdown['Name']=='Original network residual value','NPV'].tolist()[0]+cap_expense*capex_fraction['Original network']
        price_breakdown_newcompressors = price_breakdown.loc[price_breakdown['Name']=='New compressor station CAPEX','NPV'].tolist()[0]+cap_expense*capex_fraction['New compressor stations']
        price_breakdown_compressorrefurbishments = price_breakdown.loc[price_breakdown['Name']=='Compressor station refurbishment CAPEX','NPV'].tolist()[0]+cap_expense*capex_fraction['Compressor station refurbishments']
        price_breakdown_meterregulatorstationmod = price_breakdown.loc[price_breakdown['Name']=='Meter & regulator station modification CAPEX','NPV'].tolist()[0]+cap_expense*capex_fraction['Meter & regulator station modifications']
        price_breakdown_valvereplacment = price_breakdown.loc[price_breakdown['Name']=='Main-line valve replacement CAPEX','NPV'].tolist()[0]+cap_expense*capex_fraction['Main-line valve replacements']
        price_breakdown_newvalves = price_breakdown.loc[price_breakdown['Name']=='New loop-line valve CAPEX','NPV'].tolist()[0]+cap_expense*capex_fraction['New loop-line valves']
        price_breakdown_compressorfuel = price_breakdown.loc[price_breakdown['Name']=='Compressor Fuel','NPV'].tolist()[0]
        price_breakdown_inlineinspection = price_breakdown.loc[price_breakdown['Name']=='Inline inspection','NPV'].tolist()[0]
        price_breakdown_fixedOM = price_breakdown.loc[price_breakdown['Name']=='Administrative expenses','NPV'].tolist()[0] + price_breakdown.loc[price_breakdown['Name']=='Property insurance','NPV'].tolist()[0]
        price_breakdown_taxes = price_breakdown.loc[price_breakdown['Name']=='Income taxes payable','NPV'].tolist()[0]\
        - price_breakdown.loc[price_breakdown['Name'] == 'Monetized tax losses','NPV'].tolist()[0]

        if gen_inflation >0:
            price_breakdown_taxes = price_breakdown_taxes + price_breakdown.loc[price_breakdown['Name']=='Capital gains taxes payable','NPV'].tolist()[0]

        price_breakdown_check = price_breakdown_newpipe + price_breakdown_originalnetwork+price_breakdown_newcompressors+price_breakdown_compressorrefurbishments\
                              + price_breakdown_meterregulatorstationmod + price_breakdown_newvalves + price_breakdown_valvereplacment + price_breakdown_compressorfuel\
                              + price_breakdown_inlineinspection + price_breakdown_fixedOM + price_breakdown_taxes + remaining_financial
         
        price_breakdown_aggregated = {'Delivery cost: Original network ($/MMBTU)':price_breakdown_originalnetwork,\
                                      'Delivery cost: New pipe ($/MMBTU)':price_breakdown_newpipe,\
                                      'Delivery cost: New compressor stations ($/MMBTU)':price_breakdown_newcompressors,\
                                      'Delivery cost: Meter & regulator station modifications ($/MMBTU)': price_breakdown_meterregulatorstationmod,\
                                      'Delivery cost: Main-line valve replacements ($/MMBTU)': price_breakdown_valvereplacment,\
                                      'Delivery cost: New loop-line valves ($/MMBTU)': price_breakdown_newvalves,\
                                      'Delivery cost: Compressor station refurbishments ($/MMBTU)':price_breakdown_compressorrefurbishments,\
                                      'Delivery cost: Compressor fuel ($/MMBTU)':price_breakdown_compressorfuel,\
                                      'Delivery cost: Inline inspection ($/MMBTU)':price_breakdown_inlineinspection,\
                                      'Delivery cost: Fixed O&M ($/MMBTU)':price_breakdown_fixedOM,\
                                      'Delivery cost: Taxes ($/MMBTU)':price_breakdown_taxes,\
                                      'Delivery cost: Financial ($/MMBTU)':remaining_financial
                                      }

        # params_out = {'Parameter':[],'Value':[], 'Units':[]}
        params_out = []
        params_out.append(['Levelized Cost of Transport',sol['lco'],'$/MMBTU'])

        params_out.append(['LCOT: Original network',price_breakdown_originalnetwork,'$/MMBTU'])
        params_out.append(['LCOT: New pipe',price_breakdown_newpipe,'$/MMBTU'])
        params_out.append(['LCOT: Meter & regulator station modifications',price_breakdown_meterregulatorstationmod,'$/MMBTU'])
        params_out.append(['LCOT: Main-line valve replacements',price_breakdown_valvereplacment,'$/MMBTU'])
        params_out.append(['LCOT: New loop-line valves',price_breakdown_newvalves,'$/MMBTU'])
        params_out.append(['LCOT: New compressor stations',price_breakdown_newcompressors,'$/MMBTU'])
        params_out.append(['LCOT: Compressor station refurbishments',price_breakdown_compressorrefurbishments,'$/MMBTU'])
        params_out.append(['LCOT: Compressor fuel',price_breakdown_compressorfuel,'$/MMBTU'])
        params_out.append(['LCOT: Inline inspection',price_breakdown_inlineinspection,'$/MMBTU'])
        params_out.append(['LCOT: Fixed O&M',price_breakdown_fixedOM,'$/MMBTU'])
        params_out.append(['LCOT: Taxes',price_breakdown_taxes,'$/MMBTU'])
        params_out.append(['LCOT: Financial',remaining_financial,'$/MMBTU'])


        params_out.append(['Hydrogen Injection Price',H2_cost,'$/kg'])
        params_out.append(['Natural Gas Price',nat_gas_cost,'$/MMBTU'])
        params_out.append(['Blended Gas Price',cs_cost,'$/MMBTU'])
        params_out.append(['Pipeline Capacity (daily)',self.capacity,'MMBTU/day'])
        params_out.append(['Pipeline Capacity (hourly)',self.capacity/24,'MMBTU/hr'])
        params_out.append(['Added pipeline',pipeline_added,'mi'])
        params_out.append(['Added compressor stations',num_new_compressors,''])
        params_out.append(['Added compressor capacity',sum(added_CS_rating),'hp'])
        params_out.append(['Compressor fuel usage',sum(self.mod_out['CS_fuel']),'MMBTU/hr'])
        #   Capital
        params_out.append(['New pipe',total_cap,'2020$'])
        params_out.append(['New compressor stations',added_cap_compressors,'2020$']) 
        params_out.append(['Compressor station refurbishment',revamped_compressors,'2020$'])

        params_out.append(['Main-line valve replacement',pipe_main_valve_cost,'2020$'])
        params_out.append(['New loop-line valves',pipe_loop_valve_cost,'2020$'])
        params_out.append(['Meter & regulator station modifications',meter_regulator_station_modification_cost,'2020$'])
        params_out.append(['Original network residual value',original_network_residual_value,'2020$']) 

        params_out = pd.DataFrame(data=params_out,columns=['Parameter','Value','Units'])

        #   New pipe
        # Reorder columns
        
        orig_cap_costing = orig_cap_costing[['Pipe Diameter [in]','Length [mi]','Grade','Schedule']]

        #   TODO need to add names
        comp_df_out = pd.DataFrame({'Shaft Power (MW)':self.mod_out['CS_rating'],
                                    'Rated Capacity (MW)':padded_original_ratings,
                                    'Required Additional Capacity (hp)':added_CS_rating,
                                    'Number of Compressor Stations':num_comp_stations_new,
                                    'Average Compressor Station Capacity (hp)':avg_station_cap_new,
                                    'Cost (2020$)':added_cap_compressors_i,
                                    'Fuel usage (MMBTU/hr)':self.mod_out['CS_fuel']})
        

        # Write results file
        if self.design_option_original == 'no fracture control':
            design_option_towrite = 'nfc'
        else:
            design_option_towrite = self.design_option_original
        mod_type_filter = {'direct_replacement':'DR','parallel_loop':'PL','additional_compressors':'AC'}
        results_file_name = f'{self.save_directory}ResultFiles/{mod_type_filter[self.mod_type]}_{self.blending}_{design_option_towrite}_{self.batch_id}_results.xlsx'
        writer = pd.ExcelWriter(results_file_name,engine='xlsxwriter')
        workbook=writer.book

        # Write results file - disclaimer
        worksheet=workbook.add_worksheet('Disclaimer')
        startrow = 1
        worksheet.write_string(startrow, 0, disclaimer_message())
        worksheet.set_column('A:A',100,workbook.add_format({'text_wrap':True}))
        
        # Write results file - inputs sheet
        worksheet=workbook.add_worksheet('Inputs')
        writer.sheets['Inputs'] = worksheet

        inputs = {'Name':[],'Value':[]}
        inputs['Name'].extend(['Network name','Save directory','Design option - original','Location class','Joint factor',
                               'Design pressure known','Pressure drop factor','Compression ratio','Blending ratio',
                               'Natural gas cost ($/MMBTU)','H2 cost ($/kg)','Discretization nodes','Final outlet pressure (Pa)',
                               'Region','Inline inspection interval (year)', 'Batch id', 'Modification method','Modification design option'])
        inputs['Value'].extend([self.casestudy_name,self.save_directory,self.design_option_original,self.location_class,self.joint_factor,
                                self.design_pressure_known,self.pressure_drop_factor,self.compression_ratio,self.blending,
                                self.natural_gas_cost,self.hydrogen_cost,self.nodes,self.final_outlet_pressure,self.region, 
                                self.inline_inspection_interval, self.batch_id,self.mod_type, self.design_option])
        inputs['Name'].extend(financial_params.keys())
        for i in financial_params.keys():
            inputs['Value'].append(pf.vals[i.replace('_',' ')])
        inputs = pd.DataFrame(inputs)

        startrow = 1
        inputs.to_excel(writer,sheet_name='Inputs',startrow=startrow , startcol=0,index=False)


        # Write results file - results sheet
        worksheet=workbook.add_worksheet('Results')
        writer.sheets['Results'] = worksheet

        startrow = 1
        params_out.to_excel(writer,sheet_name='Results',startrow=startrow , startcol=0,index=False)
        startrow += params_out.shape[0]+4

        if self.mod_type != 'additional_compressors':
            worksheet.write_string(startrow, 0, 'Cost breakdown of new pipe')
            startrow += 1
            cap_costing_out.to_excel(writer,sheet_name='Results',startrow=startrow , startcol=0,index=False)
            startrow += cap_costing_out.shape[0]+4

        worksheet.write_string(startrow, 0, 'Breakdown of original pipe')
        startrow += 1
        orig_cap_costing.to_excel(writer,sheet_name='Results',startrow=startrow , startcol=0,index=False)
        startrow += orig_cap_costing.shape[0]+4

        worksheet.write_string(startrow, 0, 'Cost breakdown of compressors')
        startrow += 1
        comp_df_out.to_excel(writer,sheet_name='Results',startrow=startrow , startcol=0,index=False)
        startrow += comp_df_out.shape[0]+4

        # write a new sheet with new pipe designs
        if not self.mod_type=='additional_compressors':
            pipe_df_out = self.mod_out['detailed_geom']
            pipe_df_out['Length [km]'] = pipe_df_out['lengths']/1000
            pipe_df_out['Length [mi]'] = pipe_df_out['Length [km]']*0.621
            worksheet=workbook.add_worksheet('New pipe design')
            writer.sheets['New pipe design'] = worksheet

            pipe_df_out_index = pipe_df_out.reset_index()
            pipe_df_out_index = pipe_df_out_index['index']

            design_pressure_df = pd.DataFrame(self.segment_props['ASME design pressure [MPa]'])
            design_pressure_df = design_pressure_df.rename(columns = {'ASME design pressure [MPa]':'Design pressure [MPa]'})

            if self.mod_type=='parallel_loop':
                pipe_df_out = pd.concat([pipe_df_out,design_pressure_df],axis=1)


            startrow = 1
            pipe_df_out.to_excel(writer,sheet_name='New pipe design',startrow=startrow,startcol=0,index=False)
        else:
            comp_specs_df_out = self.mod_out['compressor_specs']
            worksheet = workbook.add_worksheet('New compressor design')
            writer.sheets['New compressor design'] = worksheet
            startrow = 1
            comp_specs_df_out.to_excel(writer,sheet_name='New compressor design',startrow=startrow,startcol=0,index=False)

        worksheet=workbook.add_worksheet('Demand errors')
        writer.sheets['Demand errors'] = worksheet

        startrow = 1
        demand_errors_df_out = self.mod_out['demand_errors']
        demand_errors_df_out.to_excel(writer,sheet_name = 'Demand errors',startrow=startrow,startcol=0,index=False)


        writer._save()

        return sol['lco']

    def remake_network_file(self,geom_out,mod_type):
        base_name =  f'{self.save_directory}NetworkFiles/{self.batch_id}_{self.casestudy_name}'
        network_file_in = f'{base_name}_Network_blending_in.xlsx'
        event_file_in = f'{base_name}_Event_blending_in.xlsx'
        original_GPI = pd.read_excel(network_file_in,'GPI')
        original_GNO = pd.read_excel(network_file_in,'GNO')
        original_GCS = pd.read_excel(network_file_in,'GCS')
        original_GSCE = pd.read_excel(event_file_in,'GSCE')
        gno_out_df = original_GNO
        gcs_out_df = original_GCS
        gsce_out_df = original_GSCE

        gpi_out = {'Name':[],'FromName':[],'ToName':[],'Alias':'-','InService  = True':'TRUE','Visible  = True':'TRUE','Info  = -':'-',
                    'DrawLine  = False':'FALSE','SubName':'NONE','D [mm] = 600':[],'Eff [-] = 1':1,'HTC [W/m2•K] = 0.24':0.24,
                    'L [km] = 10':[],'RO [mm] = 0.012':[],'WTH [mm] = 2':[]}

        if 'PMINDEF [MPa] = 0.101325' in gno_out_df.columns:
            gno_out_df.loc[gno_out_df['PMINDEF [MPa] = 0.101325']==0.101325,'PMINDEF [MPa] = 0.101325']=0
        if 'PMAXDEF [MPa] = 20.101325' in gno_out_df.columns:
            gno_out_df['PMAXDEF [MPa] = 20.101325'] = gno_out_df['PMAXDEF [MPa] = 20.101325'] - 0.101325
        gno_out_df=gno_out_df.rename(columns={'PMAXDEF [MPa] = 20.101325':'PMAXDEF [MPa-g] = 20.101325','PMINDEF [MPa] = 0.101325':'PMINDEF [MPa-g] = 0'})
        gno_out_df.loc[gno_out_df.index[-1],'PMINDEF [MPa-g] = 0'] = self.final_outlet_pressure/1e6
        if mod_type == 'parallel_loop':
            gno_out_df['PMAXDEF [MPa-g] = 20.101325']=self.max_node_pressure['Max Allowable Pressure [MPa]']
            for i in range(len(self.pipe_segments)): # ONLY FOR LOOPING
                pipes_in_segment = self.pipe_segments[i]
                loop_length = geom_out['lengths'][i]/1000
                loop_done = False
                if loop_length==0:
                    loop_done = True
                orig_length = 0
                for j in pipes_in_segment:
                    orig_row = original_GPI.loc[original_GPI['Name']==j]
                    this_pipe_length = orig_row["L [km] = 10"].iat[0]
                    
                    #   Go into this if it is the final segment where the looping stops
                    if (orig_length+this_pipe_length>loop_length) and (not loop_done):
                        loop_done = True
                        final_looped_length = loop_length-orig_length
                        final_pipe_length = orig_length + this_pipe_length - loop_length


                        #   Append original with 'M' modifier on the name
                        gpi_out['Name'].append(f'{orig_row["Name"].iat[0]}M')
                        gpi_out['FromName'].append(f'{orig_row["FromName"].iat[0]}')
                        gpi_out['ToName'].append(f'{orig_row["FromName"].iat[0]}a')
                        gpi_out['D [mm] = 600'].append(orig_row["D [mm] = 600"].iat[0])
                        gpi_out['L [km] = 10'].append(final_looped_length)
                        gpi_out['WTH [mm] = 2'].append(orig_row["WTH [mm] = 2"].iat[0])
                        gpi_out['RO [mm] = 0.012'].append(orig_row["RO [mm] = 0.012"].iat[0])

                        #   Append looped segment with ToName with 'a' appended
                        gpi_out['Name'].append(f'{orig_row["Name"].iat[0]}L')
                        gpi_out['FromName'].append(f'{orig_row["FromName"].iat[0]}')
                        gpi_out['ToName'].append(f'{orig_row["FromName"].iat[0]}a')
                        gpi_out['D [mm] = 600'].append(geom_out['inner diameter'][i]*1000)
                        gpi_out['L [km] = 10'].append(final_looped_length)
                        gpi_out['WTH [mm] = 2'].append(geom_out['wth'][i])
                        gpi_out['RO [mm] = 0.012'].append(orig_row["RO [mm] = 0.012"].iat[0])

                        #   Append remaining unlooped length
                        gpi_out['Name'].append(f'{orig_row["Name"].iat[0]}')
                        gpi_out['FromName'].append(f'{orig_row["FromName"].iat[0]}a')
                        gpi_out['ToName'].append(f'{orig_row["ToName"].iat[0]}')
                        gpi_out['D [mm] = 600'].append(orig_row["D [mm] = 600"].iat[0])
                        gpi_out['L [km] = 10'].append(final_pipe_length)
                        gpi_out['WTH [mm] = 2'].append(orig_row["WTH [mm] = 2"].iat[0])
                        gpi_out['RO [mm] = 0.012'].append(orig_row["RO [mm] = 0.012"].iat[0])

                        gno_out_row = gno_out_df.loc[gno_out_df['Name']==orig_row["FromName"].iat[0]].copy()
                        gno_out_row['Name'] = f'{orig_row["FromName"].iat[0]}a'
                        gno_out_df = pd.concat((gno_out_df,gno_out_row))

                    #   Otherwise add the remaining pipes as usual. If it is looped, add the looped segment as well
                    else:
                        #   Append original with 'M' modifier on the name
                        M_suffix = '' if loop_done else 'M'
                        gpi_out['Name'].append(f'{orig_row["Name"].iat[0]}{M_suffix}')
                        gpi_out['FromName'].append(f'{orig_row["FromName"].iat[0]}')
                        gpi_out['ToName'].append(f'{orig_row["ToName"].iat[0]}')
                        gpi_out['D [mm] = 600'].append(orig_row["D [mm] = 600"].iat[0])
                        gpi_out['L [km] = 10'].append(this_pipe_length)
                        gpi_out['WTH [mm] = 2'].append(orig_row["WTH [mm] = 2"].iat[0])
                        gpi_out['RO [mm] = 0.012'].append(orig_row["RO [mm] = 0.012"].iat[0])

                        #   Append looped segment with 'L' modifier on the name
                        if not loop_done:
                            gpi_out['Name'].append(f'{orig_row["Name"].iat[0]}L')
                            gpi_out['FromName'].append(f'{orig_row["FromName"].iat[0]}')
                            gpi_out['ToName'].append(f'{orig_row["ToName"].iat[0]}')
                            gpi_out['D [mm] = 600'].append(geom_out['inner diameter'][i]*1000)
                            gpi_out['L [km] = 10'].append(this_pipe_length)
                            gpi_out['WTH [mm] = 2'].append(geom_out['wth'][i])
                            gpi_out['RO [mm] = 0.012'].append(orig_row["RO [mm] = 0.012"].iat[0])

                        #   Update the looped length
                        orig_length = orig_length + orig_row["L [km] = 10"].iat[0]
        elif mod_type == 'direct_replacement':
            gno_out_df['PMAXDEF [MPa-g] = 20.101325']=self.max_node_pressure['Max Allowable Pressure [MPa]']
            pipe_count = 0
            for i in range(len(self.pipe_segments)):
                pipes_in_segment = self.pipe_segments[i]
                for p_name in pipes_in_segment:
                    orig_row = original_GPI.loc[original_GPI['Name']==p_name]
                    gpi_out['Name'].append(f'{orig_row["Name"].iat[0]}')
                    gpi_out['FromName'].append(f'{orig_row["FromName"].iat[0]}')
                    gpi_out['ToName'].append(f'{orig_row["ToName"].iat[0]}')
                    gpi_out['D [mm] = 600'].append(geom_out['outer diameter'][pipe_count]- 2*geom_out['wth'][pipe_count])
                    gpi_out['L [km] = 10'].append(orig_row["L [km] = 10"].iat[0])
                    gpi_out['WTH [mm] = 2'].append(geom_out['wth'][pipe_count])
                    gpi_out['RO [mm] = 0.012'].append(orig_row["RO [mm] = 0.012"].iat[0])
                    pipe_count+=1
        elif mod_type == 'additional_compressors':
            compression_ratio_by_segment = geom_out['compression_ratio']
            cum_comp_length = [np.cumsum(x)/1000 for x in geom_out['compressor_length']]
            for i in range(len(self.pipe_segments)):
                pipes_in_segment = self.pipe_segments[i]
                seg_pmax = self.segment_props['ASME design pressure [MPa]'].iat[0]
                total_length = 0
                prev_length = 0
                min_comp_index = 0
                for j in pipes_in_segment:
                    orig_row = original_GPI.loc[original_GPI['Name']==j]
                    this_pipe_length = orig_row["L [km] = 10"].iat[0]
                    prev_length = total_length
                    total_length = total_length + this_pipe_length
                    comp_added = False
                    c_naming_index = min_comp_index

                    gno_out_df.loc[gno_out_df['Name']==orig_row['FromName'].iat[0],'PMAXDEF [MPa-g] = 20.101325'] = seg_pmax
                    gno_out_df.loc[gno_out_df['Name']==orig_row['ToName'].iat[0],'PMAXDEF [MPa-g] = 20.101325'] = seg_pmax

                    for ci,c in enumerate(cum_comp_length[i]):
                        if ci<min_comp_index:
                            continue
                        if c<total_length:
                            pipe_subscript = chr(97+ci-c_naming_index)
                            node_prev = chr(97+ci-c_naming_index-1) if comp_added else ''
                            gpi_out['Name'].append(f'{orig_row["Name"].iat[0]}{pipe_subscript}')
                            gpi_out['FromName'].append(f'{orig_row["FromName"].iat[0]}{node_prev}')
                            gpi_out['ToName'].append(f'{orig_row["FromName"].iat[0]}{pipe_subscript}_c')
                            gpi_out['D [mm] = 600'].append(orig_row["D [mm] = 600"].iat[0])
                            gpi_out['L [km] = 10'].append(c-prev_length)
                            gpi_out['WTH [mm] = 2'].append(orig_row["WTH [mm] = 2"].iat[0])
                            gpi_out['RO [mm] = 0.012'].append(orig_row["RO [mm] = 0.012"].iat[0])
                            comp_added = True
                            min_comp_index = min_comp_index+1
                            prev_length = c

                            gno_out_row = gno_out_df.loc[gno_out_df['Name']==orig_row["FromName"].iat[0]].copy()
                            gno_out_row['Name'] = f'{orig_row["FromName"].iat[0]}{pipe_subscript}_c'
                            gno_out_row['PMAXDEF [MPa-g] = 20.101325'] = seg_pmax
                            gno_out_df = pd.concat((gno_out_df,gno_out_row))
                            gno_out_row['Name'] = f'{orig_row["FromName"].iat[0]}{pipe_subscript}'
                            gno_out_row['PMAXDEF [MPa-g] = 20.101325'] = seg_pmax
                            gno_out_df = pd.concat((gno_out_df,gno_out_row))
                            

                            if gcs_out_df.empty:
                                gcs_out_row = gcs_out_df.copy()
                                gcs_out_row.loc[1] = [None]*len(gcs_out_df.columns)
                                gcs_out_row['Alias'] = '-'
                                gcs_out_row['InService  = True'] = 1
                                gcs_out_row['Visible  = True'] = 1
                                gcs_out_row['Info  = -'] = '-'
                                gcs_out_row['DrawLine  = False'] = 0
                                gcs_out_row['SubName'] = 'SUB1'
                                gcs_out_row['CPTYPE  = Centrifugal'] = 'Centrifugal'
                                gcs_out_row['DRVTYPE  = Gas'] = 'Gas'
                                gcs_out_row['D [mm] = 600'] = 600
                                gcs_out_row['EFFHDEF [-] = 0.87'] = 0.87
                                gcs_out_row['EFFMDEF [-] = 0.34'] = 0.357
                                gcs_out_row['PIMINDEF [MPa] = 0.101325'] = 0.101325
                                gcs_out_row['PIMINPRCDEF [$/bar] = 100000000'] = 100000000
                                gcs_out_row['POMAXDEF [MPa-g] = 20.101325'] = seg_pmax
                                gcs_out_row['POMAXPRCDEF [$/bar] = ∞'] = '∞'
                                gcs_out_row['POWDMAXPRCDEF [€/MWh] = ∞'] = '∞'
                                gcs_out_row['POWSMAXPRCDEF [€/MWh] = ∞'] = '∞'
                                gcs_out_row['PRMAXPRCDEF [€] = ∞'] = '∞'
                                gcs_out_row['QMAXDEF [ksm3/h] = 3600'] = 10000
                                gcs_out_row['QMAXPRCDEF [$/sm3] = ∞'] = '∞'
                                gcs_out_row['QVOLMAXDEF [m3/s] = 100'] = 10000
                                gcs_out_row['QVOLMAXPRCDEF [$/m3] = 1000'] = 1000
                                gcs_out_row['VMAXDEF [m/s] = 60'] = 60
                                gcs_out_row['VMAXPRCDEF [$/m] = 1000'] = 1000

                            else:
                                gcs_out_row = gcs_out_df.iloc[:1].copy()
                            gcs_out_row['Name'] = f'BS{i+1}{chr(97+ci)}'
                            gcs_out_row['FromName'] = f'{orig_row["FromName"].iat[0]}{pipe_subscript}_c'
                            gcs_out_row['ToName'] = f'{orig_row["FromName"].iat[0]}{pipe_subscript}'
                            gcs_out_row['PRMAXDEF [-] = 3'] = 3
                            gcs_out_row['POWDMAXDEF [MW] = 200'] = 700
                            gcs_out_row['POWSMAXDEF [MW] = 100'] = 350

                            # gcs_out_row['POMAXDEF [MPa-g] = 20'] = node_pmax
                            gcs_out_df = pd.concat((gcs_out_df,gcs_out_row))

                            comp_setting = 'PRSET' if ci==len(cum_comp_length[i])-1 else 'POSET'
                            gsce_value = compression_ratio_by_segment[i][ci] if ci==len(cum_comp_length[i])-1 else seg_pmax
                            gsce_row = pd.DataFrame({'Parameter':f'GCS.BS{i+1}{chr(97+ci)}.{comp_setting}','Value':gsce_value,'Unit':'[MPa-g]','Active':'True','Info ':'-'},index=[0])
                            gsce_out_df = pd.concat((gsce_out_df,gsce_row))


                    if not comp_added:
                        gpi_out['Name'].append(f'{orig_row["Name"].iat[0]}')
                        gpi_out['FromName'].append(f'{orig_row["FromName"].iat[0]}')
                        gpi_out['ToName'].append(f'{orig_row["ToName"].iat[0]}')
                        gpi_out['D [mm] = 600'].append(orig_row["D [mm] = 600"].iat[0])
                        gpi_out['L [km] = 10'].append(this_pipe_length)
                        gpi_out['WTH [mm] = 2'].append(orig_row["WTH [mm] = 2"].iat[0])
                        gpi_out['RO [mm] = 0.012'].append(orig_row["RO [mm] = 0.012"].iat[0])
                        
                    else:
                        gpi_out['Name'].append(f'{orig_row["Name"].iat[0]}')
                        gpi_out['FromName'].append(f'{orig_row["FromName"].iat[0]}{pipe_subscript}')
                        gpi_out['ToName'].append(f'{orig_row["ToName"].iat[0]}')
                        gpi_out['D [mm] = 600'].append(orig_row["D [mm] = 600"].iat[0])
                        gpi_out['L [km] = 10'].append(total_length-prev_length)
                        gpi_out['WTH [mm] = 2'].append(orig_row["WTH [mm] = 2"].iat[0])
                        gpi_out['RO [mm] = 0.012'].append(orig_row["RO [mm] = 0.012"].iat[0])
            gcs_out_df['ExtractFuel  = False'] = 1
        gpi_out_df = pd.DataFrame(gpi_out)
        
        #   Copy the files the to the new directory before editting
        base_name =  f'{self.save_directory}NetworkFiles/{self.batch_id}_{self.casestudy_name}'
        new_file_name = f'{base_name}_Network_new.xlsx'
        new_file_name_event = f'{base_name}_Event_new.xlsx'
        shutil.copy(f'{base_name}_Network_blending_in.xlsx',new_file_name)
        shutil.copy(f'{base_name}_Event_blending_in.xlsx',new_file_name_event)

        #   Edit the max pressure for compressors in the network file
        gcs_out_df=gcs_out_df.rename(columns={'POMAXDEF [MPa] = 20.101325':'POMAXDEF [MPa-g] = 20.101325'})
        gcs_out_df['POMAXDEF [MPa-g] = 20.101325'] = (gcs_out_df.merge(gno_out_df[['Name','PMAXDEF [MPa-g] = 20.101325']],left_on='ToName',right_on='Name'))['PMAXDEF [MPa-g] = 20.101325'].values
        

        #   Edit the pressure in the event file 
        if mod_type in ['parallel_loop','additional_compressors', 'direct_replacement']:

            gsce_gsup_col = [col for col in gsce_out_df['Parameter'].values if (('GSUP' in col) and ('PSET' in col))][0]
            gsup_node = gsce_gsup_col.split('.')[1]
            gsce_out_df.loc[gsce_out_df['Parameter']==gsce_gsup_col,'Value'] = self.max_node_pressure.loc[self.max_node_pressure['Node Name']==gsup_node,'Max Allowable Pressure [MPa]'].iat[0]
            gsce_out_df.loc[gsce_out_df['Parameter']==gsce_gsup_col,'Unit'] = '[MPa-g]'
            gsce_gsup_col = [col for col in gsce_out_df['Parameter'].values if (('GCS' in col) and ('POSET' in col) and ('BS' not in col) )]
            
            for col in gsce_gsup_col:
                gcs_name = col.split('.')[1]
                gsce_out_df.loc[gsce_out_df['Parameter']==col,'Value'] = gcs_out_df.loc[gcs_out_df['Name']==gcs_name,'POMAXDEF [MPa-g] = 20.101325'].iat[0]
                gsce_out_df.loc[gsce_out_df['Parameter']==col,'Unit'] = '[MPa-g]'

        #   First delete previous sheets. Because overwrite option only exists in pandas>=1.3
        workbook1 = openpyxl.load_workbook(new_file_name)
        del workbook1['GPI'],workbook1['GNO'],workbook1['GCS']
        workbook1.save(new_file_name)
        workbook1 = openpyxl.load_workbook(new_file_name_event)
        del workbook1['GSCE']
        workbook1.save(new_file_name_event)
        #   Write to the file
        with pd.ExcelWriter(new_file_name,engine='openpyxl',mode="a",if_sheet_exists='replace') as writer:  
            gpi_out_df.to_excel(writer, sheet_name='GPI',index=False)
            gno_out_df.sort_values(by='Name').to_excel(writer, sheet_name='GNO',index=False)
            gcs_out_df.to_excel(writer, sheet_name='GCS',index=False)
        with pd.ExcelWriter(new_file_name_event,engine='openpyxl',mode="a",if_sheet_exists='replace') as writer:  
            gsce_out_df.to_excel(writer, sheet_name='GSCE',index=False)

        return gcs_out_df['Name'].values
    
    def runModifiedFile(self,mod_type,CS_names):
        '''
            Run SAInt using modified files
        '''
        mydll_new = cdll.LoadLibrary(self.dllfolder + "SAInt-API.dll")
        mydll_new.evalStr.restype=c_wchar_p
        mydll_new.evalInt.restype=c_int
        mydll_new.evalCmdInt.restype=c_int
        mydll_new.evalBool.restype=c_bool
        mydll_new.evalFloat.restype=c_float
        mydll_new, status = daf.runSAInt(mydll_new,f'{self.save_directory}NetworkFiles/{self.batch_id}_{self.casestudy_name}','_new')

        CS_fuel,CS_rating,comp_params = daf.get_compressor_usage(mydll_new,CS_names) # MMBTU/hr for CS_fuel
        CS_PR = comp_params['Pressure ratio [-]'].tolist()

        if mod_type == 'additional_compressors':
            comp_locations = daf.compressor_locations(f'{self.save_directory}NetworkFiles/',f'{self.batch_id}_{self.casestudy_name}'+'_Network_new.xlsx')

            compressor_specs = pd.merge(comp_locations,comp_params,how = 'left',left_on='Name',right_on='Name')
            compressor_specs['Compressor station rated capacity [HP]'] = compressor_specs['Compressor Rating [MW]']*1341.02
        else:
            compressor_specs = 0

        # Get demand constraints from edited model
        demand_constraints = daf.get_demand_constraints(f'{self.save_directory}NetworkFiles/',f'{self.batch_id}_{self.casestudy_name}')
        
        # Get demand flows to compare to constraints
        demand_flows = daf.get_demand_flows(mydll_new,f'{self.save_directory}NetworkFiles/',f'{self.batch_id}_{self.casestudy_name}')

        # Calculate demand errors
        demand_errors = demand_constraints.rename(columns = {'Value':'Demand flow targets [MW]'})
        demand_errors['Actual demand flows [MW]'] = demand_flows['Value']
        demand_errors['Error (%)'] = (demand_errors['Demand flow targets [MW]']-demand_errors['Actual demand flows [MW]'])/(demand_errors['Demand flow targets [MW]']+1e-6)*100


        return {'CS_fuel':CS_fuel,'CS_rating':CS_rating,'CS_PR':CS_PR,'compressor_specs':compressor_specs,'demand_errors':demand_errors, 'Status': status}